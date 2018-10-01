from openpyxl import load_workbook
wb2 = load_workbook('Final.xlsx')
print (wb2.sheetnames)
#ws = wb2.active
ws =wb2["FULL"]
c = ws['A2']
print(c.value)
# Function to find the column of Duns
duns_col = ''
for row in ws.rows:
    for cell in row:
        if(cell.value == "D-U-N-S® Number"):
            duns_col = cell.column
            print(duns_col)
            break
    break
# Count zero duns and non zero duns    
row = 1
countnzero = 0
countzero = 0
rows = 1
for row in ws.rows:
    rows = rows + 1

print("Rows: "+str(rows-1))    

for row in range(2,rows):        
    dun = ws[duns_col+str(row)].value    
    if (dun == "000000000"):
        countzero = countzero +1
    else:
        countnzero = countnzero +1
    row = row + 1    

print("Unmatched :"+str(countzero))
print("Matched :"+str(countnzero))



       
